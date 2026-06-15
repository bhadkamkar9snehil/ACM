from __future__ import annotations

import csv
import os
import sys
import shutil
from pathlib import Path
from typing import Any
from sim.models import CsvMetadata, CsvFileRecord, CsvPreviewResponse, TagMapping
from sim.type_inference import infer_types, sanitize_tag_name, is_default_disabled_column

ROOT = Path(__file__).resolve().parents[1]
UPLOAD_DIR = ROOT / "sim_data" / "uploads"
GENERATED_DIR = ROOT / "sim_data" / "generated"
SAMPLE_DIR = ROOT / "sim_data" / "sample"
SIM_DATA_DIR = ROOT / "sim_data"
SOURCE_DIRS = {"uploaded": UPLOAD_DIR, "generated": GENERATED_DIR, "sample": SAMPLE_DIR}


def ensure_dirs() -> None:
    for path in SOURCE_DIRS.values():
        path.mkdir(parents=True, exist_ok=True)


def sanitize_filename(filename: str) -> str:
    raw = filename or ""
    if not raw or raw in {".", ".."} or ".." in raw or "/" in raw or "\\" in raw or "\x00" in raw:
        raise ValueError("Unsafe filename.")
    base = os.path.basename(raw)
    if not base.lower().endswith((".csv", ".xlsx")):
        raise ValueError("Filename must end with .csv or .xlsx.")
    return base


def resolve_csv_path(filename: str, source: str) -> Path:
    if source not in SOURCE_DIRS:
        raise ValueError("Invalid CSV source.")
    safe = sanitize_filename(filename)
    path = SOURCE_DIRS[source] / safe
    root = SOURCE_DIRS[source].resolve()
    resolved = path.resolve()
    if root not in resolved.parents and resolved != root:
        raise ValueError("Unsafe CSV path.")
    if not resolved.exists():
        raise FileNotFoundError(f"CSV file not found: {safe}")
    return resolved


def read_rows(path: Path, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]]]:
    if path.suffix.lower() == ".xlsx":
        return read_xlsx_rows(path, max_rows=max_rows)
    
    import pandas as pd
    try:
        df = pd.read_csv(str(path), sep=None, engine="python", nrows=max_rows)
    except pd.errors.EmptyDataError:
        raise ValueError("File has no header.")
        
    df = df.fillna("")
    columns = [str(c).strip() for c in df.columns]
    if len(set(columns)) != len(columns):
        raise ValueError("File has duplicate columns.")
        
    df.columns = columns
    rows = df.astype(str).to_dict(orient="records")
    return columns, rows


def read_xlsx_rows(path: Path, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError("openpyxl is required to read Excel .xlsx files.") from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration:
        raise ValueError("Excel file has no header row.")
    columns = [str(c).strip() if c is not None else "" for c in header]
    if not any(columns):
        raise ValueError("Excel file has no header row.")
    if any(not c for c in columns):
        raise ValueError("Excel header contains blank column names.")
    if len(set(columns)) != len(columns):
        raise ValueError("Excel file has duplicate columns.")
    rows: list[dict[str, str]] = []
    for idx, values in enumerate(iterator):
        if max_rows is not None and idx >= max_rows:
            break
        row: dict[str, str] = {}
        for col, value in zip(columns, values):
            row[col] = "" if value is None else str(value)
        for col in columns[len(values or ()):]:
            row[col] = ""
        if any(v != "" for v in row.values()):
            rows.append(row)
    return columns, rows


def read_full_csv(filename: str, source: str, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]]]:
    return read_rows(resolve_csv_path(filename, source), max_rows=max_rows)


def metadata(filename: str, source: str, preview_rows: int = 10) -> CsvMetadata:
    path = resolve_csv_path(filename, source)
    columns, rows = read_rows(path, max_rows=preview_rows)
    inferred = infer_types(rows, columns)
    mappings = default_tag_mappings(columns, inferred)
    
    _, row_count = _fast_count(path)
    
    return CsvMetadata(
        filename=path.name,
        source=source,  # type: ignore[arg-type]
        row_count=row_count,
        column_count=len(columns),
        columns=columns,
        preview=rows,
        inferred_types=inferred,  # type: ignore[arg-type]
        modified_at=_mtime(path),
        default_tag_mappings=mappings,
    )


def preview(filename: str, source: str, limit: int = 10) -> CsvPreviewResponse:
    path = resolve_csv_path(filename, source)
    columns, rows = read_rows(path, max_rows=limit)
    return CsvPreviewResponse(filename=path.name, source=source, columns=columns, rows=rows)  # type: ignore[arg-type]


def save_upload(file: Any) -> CsvMetadata:
    ensure_dirs()
    safe = sanitize_filename(file.filename or "upload.csv")
    target = UPLOAD_DIR / safe
    with target.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    return metadata(safe, "uploaded")


def write_rows(filename: str, rows: list[dict[str, Any]], source: str = "generated") -> Path:
    ensure_dirs()
    if source not in SOURCE_DIRS:
        raise ValueError("Invalid output source.")
    safe = sanitize_filename(filename)
    if not rows:
        raise ValueError("No rows generated.")
    path = SOURCE_DIRS[source] / safe
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _fast_count(path: Path) -> tuple[int, int]:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        return ws.max_column, max(0, ws.max_row - 1)
    
    import pandas as pd
    try:
        df_cols = pd.read_csv(str(path), sep=None, engine="python", nrows=0)
        col_count = len(df_cols.columns)
    except Exception:
        col_count = 0
        
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            next(reader)
        except StopIteration:
            pass
        return col_count, sum(1 for _ in reader)

def list_files() -> list[CsvFileRecord]:
    ensure_dirs()
    records: list[CsvFileRecord] = []
    for source, directory in SOURCE_DIRS.items():
        for path in sorted(list(directory.glob("*.csv")) + list(directory.glob("*.xlsx"))):
            try:
                col_count, row_count = _fast_count(path)
                records.append(CsvFileRecord(
                    filename=path.name,
                    source=source,  # type: ignore[arg-type]
                    path=str(path.relative_to(ROOT)),
                    row_count=row_count,
                    column_count=col_count,
                    modified_at=_mtime(path),
                ))
            except Exception:
                continue
    return records


def delete_file(filename: str, source: str) -> dict[str, bool]:
    if source == "sample":
        raise ValueError("Sample files cannot be deleted.")
    path = resolve_csv_path(filename, source)
    path.unlink()
    return {"deleted": True}


def default_tag_mappings(columns: list[str], inferred_types: dict[str, str], prefix: str = "TagSimulator") -> list[TagMapping]:
    mappings: list[TagMapping] = []
    for col in columns:
        tag = sanitize_tag_name(col)
        mappings.append(TagMapping(
            enabled=not is_default_disabled_column(col),
            csv_column=col,
            tag_name=tag,
            node_id=f"{prefix}.{tag}",
            data_type=inferred_types.get(col, "String"),  # type: ignore[arg-type]
            initial_value=None,
            writable=False,
        ))
    return mappings


def _mtime(path: Path) -> str:
    from datetime import datetime, timezone
    return datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat().replace("+00:00", "Z")
